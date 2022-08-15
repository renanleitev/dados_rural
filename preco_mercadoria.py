"""
DADOS OBTIDOS DO SITE SCOTCONSULTORIA: https://www.scotconsultoria.com.br/
PROJETO SEM FINS LUCRATIVOS, APENAS PARA FINS EDUCACIONAIS
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from datetime import datetime


options = webdriver.ChromeOptions()
options.add_argument("--start-maximized")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), chrome_options=options) # noqa

dia = datetime.now().day
mes = datetime.now().month
ano = datetime.now().year
data = f'{dia}-{mes}-{ano}'


class Driver():

    def iniciar(self):
        global driver

    def cookies(self):
        cookie = driver.find_element(By.LINK_TEXT, 'Aceitar Cookies')
        cookie.click()

    def boi_gordo(self, site):
        driver.get(site)
        boi = driver.find_element(By.XPATH, '//*[@id="geral"]/div[2]/div[2]/div/table/tbody') # noqa
        funrural = driver.find_element(By.XPATH, '//*[@id="geral"]/div[2]/div[2]/div/table/thead/tr/th[1]')  # noqa
        senar = driver.find_element(By.XPATH, '//*[@id="geral"]/div[2]/div[2]/div/table/thead/tr/th[2]') # noqa
        with open('lista_boi.txt', 'w', encoding='utf-8') as file:
            file.write(boi.text)
        lista_boi = []
        with open('lista_boi.txt', 'r', encoding='utf-8') as file:
            lista_boi.extend(iter(file))
        book = Workbook()
        sheet = book.active
        sheet.merge_cells('A1:D1')
        sheet['A1'].value = funrural.text
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('E1:G1')
        sheet['E1'].value = senar.text
        sheet['E1'].alignment = Alignment(horizontal='center')
        sheet["A2"] = 'Boi Gordo'
        sheet.merge_cells('B2:G2')
        sheet['B2'].value = 'R$/@ - Kg'
        sheet['B2'].alignment = Alignment(horizontal='center')
        sheet["A3"] = "Estado"
        sheet["B3"] = "Cidade/Região"
        sheet["C3"] = "À vista"
        sheet["D3"] = "30 dias"
        sheet["E3"] = "'%' em relação a SP"
        sheet["F3"] = "À vista"
        sheet["G3"] = "30 dias"
        with open('lista_boi.txt', 'r', encoding='utf-8') as file:
            lista_boi = []
            lista_temp = []
            for linha in file:
                lista_temp.append(linha)
                listagem = str(lista_temp).replace('[', '').replace(']', '').replace("\\n'", '').replace("'", '').replace('*', '') # noqa
                lista = list(listagem.split(" "))
                lista_boi.append(lista)
                lista_temp = []
            for c in lista_boi:
                if len(c) == 8:
                    ret_val = c.pop(2)
                    c[1] = f'{c[1]} {ret_val}'
                elif len(c) == 6:
                    c.insert(1, 'Sem cidade')
        colunas = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
        for conta, linha in enumerate(lista_boi[2:], start=4):
            sheet[f"A{conta}"] = linha[0]
            sheet[f"B{conta}"] = linha[1]
            for i in range(2, 7):
                try:
                    sheet[f"{colunas[i]}{conta}"] = float(linha[i].replace(',', '.')) # noqa
                except Exception:
                    sheet[f"{colunas[i]}{conta}"] = linha[i]
                if i == 4:
                    try:
                        sheet[f"{colunas[i]}{conta}"] = float(linha[i].replace(',', '.').replace('%', '')) # noqa
                    except Exception:
                        sheet[f"{colunas[i]}{conta}"] = linha[i]
            conta += 1
        book.save(f'lista_boi_gordo_{data}.xlsx')
        book.close()

    def vaca_gorda(self, site):
        driver.get(site)
        boi = driver.find_element(By.XPATH, '//*[@id="geral"]/div[2]/div[2]/div/table/tbody') # noqa
        funrural = driver.find_element(By.XPATH, '//*[@id="geral"]/div[2]/div[2]/div/table/thead/tr/th[1]')  # noqa
        senar = driver.find_element(By.XPATH, '//*[@id="geral"]/div[2]/div[2]/div/table/thead/tr/th[2]') # noqa
        with open('lista_boi.txt', 'w', encoding='utf-8') as file:
            file.write(boi.text)
        lista_boi = []
        with open('lista_boi.txt', 'r', encoding='utf-8') as file:
            lista_boi.extend(iter(file))
        book = Workbook()
        sheet = book.active
        sheet.merge_cells('A1:D1')
        sheet['A1'].value = funrural.text
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('E1:F1')
        sheet['E1'].value = senar.text
        sheet['E1'].alignment = Alignment(horizontal='center')
        sheet["A2"] = 'Vaca Gorda'
        sheet.merge_cells('B2:F2')
        sheet['B2'].value = 'R$/@ - Kg'
        sheet['B2'].alignment = Alignment(horizontal='center')
        sheet["A3"] = "Estado"
        sheet["B3"] = "Cidade/Região"
        sheet["C3"] = "À vista"
        sheet["D3"] = "30 dias"
        sheet["E3"] = "À vista"
        sheet["F3"] = "30 dias"
        with open('lista_boi.txt', 'r', encoding='utf-8') as file:
            lista_boi = []
            lista_temp = []
            for linha in file:
                lista_temp.append(linha)
                listagem = str(lista_temp).replace('[', '').replace(']', '').replace("\\n'", '').replace("'", '').replace('*', '') # noqa
                lista = list(listagem.split(" "))
                lista_boi.append(lista)
                lista_temp = []
            for c in lista_boi:
                if len(c) == 7:
                    ret_val = c.pop(2)
                    c[1] = f'{c[1]} {ret_val}'
                elif len(c) == 5:
                    c.insert(1, 'Sem cidade')
        colunas = ['A', 'B', 'C', 'D', 'E', 'F']
        for conta, linha in enumerate(lista_boi[2:], start=4):
            sheet[f"A{conta}"] = linha[0]
            sheet[f"B{conta}"] = linha[1]
            for i in range(2, 6):
                try:
                    sheet[f"{colunas[i]}{conta}"] = float(linha[i].replace(',', '.')) # noqa
                except Exception:
                    sheet[f"{colunas[i]}{conta}"] = linha[i]
            conta += 1
        book.save(f'lista_vaca_gorda_{data}.xlsx')
        book.close()

    def novilha(self, site):
        driver.get(site)
        boi = driver.find_element(By.XPATH, '//*[@id="geral"]/div[2]/div[2]/div/table/tbody') # noqa
        funrural = driver.find_element(By.XPATH, '//*[@id="geral"]/div[2]/div[2]/div/table/thead/tr/th[1]')  # noqa
        senar = driver.find_element(By.XPATH, '//*[@id="geral"]/div[2]/div[2]/div/table/thead/tr/th[2]') # noqa
        with open('lista_boi.txt', 'w', encoding='utf-8') as file:
            file.write(boi.text)
        lista_boi = []
        with open('lista_boi.txt', 'r', encoding='utf-8') as file:
            lista_boi.extend(iter(file))
        book = Workbook()
        sheet = book.active
        sheet.merge_cells('A1:D1')
        sheet['A1'].value = funrural.text
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet.merge_cells('E1:F1')
        sheet['E1'].value = senar.text
        sheet['E1'].alignment = Alignment(horizontal='center')
        sheet["A2"] = 'Novilha Gorda'
        sheet.merge_cells('B2:F2')
        sheet['B2'].value = 'R$/@ - Kg'
        sheet['B2'].alignment = Alignment(horizontal='center')
        sheet["A3"] = "Estado"
        sheet["B3"] = "Cidade/Região"
        sheet["C3"] = "À vista"
        sheet["D3"] = "30 dias"
        sheet["E3"] = "À vista"
        sheet["F3"] = "30 dias"
        with open('lista_boi.txt', 'r', encoding='utf-8') as file:
            lista_boi = []
            lista_temp = []
            for linha in file:
                lista_temp.append(linha)
                listagem = str(lista_temp).replace('[', '').replace(']', '').replace("\\n'", '').replace("'", '').replace('*', '') # noqa
                lista = list(listagem.split(" "))
                lista_boi.append(lista)
                lista_temp = []
            for c in lista_boi:
                if len(c) == 7:
                    ret_val = c.pop(2)
                    c[1] = f'{c[1]} {ret_val}'
                elif len(c) == 5:
                    c.insert(1, 'Sem cidade')
        colunas = ['A', 'B', 'C', 'D', 'E', 'F']
        for conta, linha in enumerate(lista_boi[2:], start=4):
            sheet[f"A{conta}"] = linha[0]
            sheet[f"B{conta}"] = linha[1]
            for i in range(2, 6):
                try:
                    sheet[f"{colunas[i]}{conta}"] = float(linha[i].replace(',', '.')) # noqa
                except Exception:
                    sheet[f"{colunas[i]}{conta}"] = linha[i]
        book.save(f'lista_novilha_{data}.xlsx')
        book.close()


if __name__ == "__main__":
    nav = Driver()
    nav.boi_gordo("https://www.scotconsultoria.com.br/cotacoes/boi-gordo/?ref=smn") # noqa
    nav.cookies()
    nav.vaca_gorda("https://www.scotconsultoria.com.br/cotacoes/vaca-gorda/?ref=smn") # noqa
    nav.novilha("https://www.scotconsultoria.com.br/cotacoes/novilha/?ref=smn") # noqa
    driver.close()
    print('Excel salvo.')
